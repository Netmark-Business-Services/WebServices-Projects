from datetime import datetime
import os
import requests
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from requests.auth import HTTPBasicAuth
import pandas as pd

# Load the environment variable
current_env = os.getenv('ENV', 'UAT')  # Default to 'Dev' if the environment variable is not set

# Define environment configurations
environments = {
    "UAT": {
        "url": "https://gchp-con-ut01.gchp.local:9193/connector/services/v4/EnrollmentSparse",
        "member_url": "https://gchp-con-ut01.gchp.local:9193/connector/services/v4/MembershipSparseLookup",
        "log_file": "",
        "summary_file": ""
    },
    "Prod": {
        "url": "https://gchp-con-pr01.gchp.local:9193/connector/services/v4/EnrollmentSparse",
        "member_url": "https://gchp-con-ut01.gchp.local:9193/connector/services/v4/MembershipSparseLookup",
        "log_file": "",
        "summary_file": ""
    },
    "PreProd": {
        "url": "https://gchp-con-pp01.gchp.local:9193/connector/services/v4/EnrollmentSparse",
        "log_file": ",
        "summary_file": ""
    },
    "Dev": {
        "url": "https://gchp-con-dv01.gchp.local:9193/connector/services/v4/EnrollmentSparse",
        "username": "connector",
        "password": "Connector123",
        "log_file": "",
        "summary_file": ""
    }
}

# Ensure the specified environment is valid
if current_env not in environments:
    raise ValueError(f"Unknown environment: {current_env}")

# Get the configuration for the current environment
config = environments[current_env]
url = config["url"]
log_file = config["log_file"]
summary_file = config["summary_file"]

# Load the Excel file
excel_file = ''
# Check if the file exists
if not os.path.exists(excel_file):
    print(f"File not found: {excel_file}")
    raise FileNotFoundError(f"The file does not exist in the specified path.")

workbook = load_workbook(filename=excel_file)
sheet = workbook.active

# Initialize a list to collect response data for the summary
response_data = []

# Initialize a counter for the number of rows processed
rows_processed = 0

# Function to remove leading zeros from the vendor ID
def clean_vendor_id(vendor_id):
    if isinstance(vendor_id, str) and vendor_id.startswith('000'):
        return vendor_id.lstrip('0')  # Strip leading zeros
    return vendor_id  # Return the ID unchanged if no leading zeros are found


    # Send the SOAP request with basic authentication
    headers = {'Content-Type': 'text/xml'}
    response = requests.post(url, data=soap_body, headers=headers, verify=False)


# Check if the summary file already exists
summary_exists = os.path.exists(summary_file)
  
# Open a file to write the responses for the current environment
with open(log_file, 'w') as response_file:
    # Iterate through the first 10 rows in the Excel file, assuming HCC IDs are in the first column
    # Remove max_row=11 to process the full file.
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row,  values_only=True): 
        hcc_id = row[0]  # Subscriber ID is the HCC ID
        vendor_id = row[1]  # Vendor ID
        pcp_name = row[2]  # PCP Name
        start_date = row[3]  # Start Date
        end_date = row[4] # END Date

        # Print the vendor ID and start date before passing them into the XML request
        print(f"HCC ID: {hcc_id}, Vendor ID: {vendor_id}, PCP Name: {pcp_name}, Start Date: {start_date}, End Date:{end_date}")
        
        # Convert the date format from DD-MM-YYYY to YYYY-MM-DD
        start_date_converted = start_date.strftime('%Y-%m-%d')
        
        # Convert the date format from DD-MM-YYYY to YYYY-MM-DD
        end_date_converted = end_date.strftime('%Y-%m-%d')

        # Clean the vendor ID by removing leading zeros if applicable
        vendor_id = clean_vendor_id(vendor_id)
        
        # Remove the '01' suffix if present
        if isinstance(hcc_id, str) and hcc_id.endswith('01'):
            hcc_id = hcc_id[:-2]

        # Create the SOAP request body
        soap_body = f"""
        <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:enr="http://www.healthedge.com/connector/schema/enrollmentsparse">
            <soapenv:Body>
                <enr:enrollment>
                    <asOfDate>Date                                                                                                                                                                                                                                                                                                                                                                                                                                                   </asOfDate>
                    <sendToWorkBasketIfExceptionsPresent>false</sendToWorkBasketIfExceptionsPresent>
                    <member>
                        <maintenanceTypeCode>CHANGE</maintenanceTypeCode>
                        <hccIdentifier>{hcc_id}</hccIdentifier>
                        <memberMatchData>
                            <definitionName>MemberMatchDefinition</definitionName>
                            <id>{hcc_id}</id>
                        </memberMatchData>
                        <providerSelections>
                            <providerSelection>
                                <providerRoleType>PCP</providerRoleType>
                                <providerDateRanges>
                                    <startDate>{start_date_converted}</startDate>
                                    <endDate>{end_date_converted}</endDate>
                                    <providerMatch>
                                        <supplierLocation>
                                            <hccIdentificationNumber>{vendor_id}</hccIdentificationNumber>
                                        </supplierLocation>
                                    </providerMatch>
                                    <pcpAutoAssigned>false</pcpAutoAssigned>
                                </providerDateRanges>
                            </providerSelection>
                        </providerSelections>
                    </member>         
                </enr:enrollment>
            </soapenv:Body>
        </soapenv:Envelope>
        """

        # Send the SOAP request with basic authentication
        headers = {'Content-Type': 'text/xml'}
        response = requests.post(url, data=soap_body, headers=headers, verify=False)

        # Write the response to the file
        response_file.write(f"Response for HCC ID {hcc_id}:\n")
        response_file.write(response.text + "\n\n")

        # Print the response (for debugging)
        print(f"Response for HCC ID {hcc_id} in {current_env}:")
        print(response.text)

       # Parse the response to extract needed information
        root = ET.fromstring(response.text)

        status = root.find(".//status")
        status_text = status.text if status is not None else 'Unknown'

        cvc_id = root.find(".//cvcId")
        cvc_id_text = cvc_id.text if cvc_id is not None else 'Unknown'

        error_message = root.find(".//messageDescription")
        error_message_text = error_message.text if error_message is not None else 'None'

        # Collect response data
        response_data.append([hcc_id, vendor_id, status_text, cvc_id_text, error_message_text])
        
        # Create a DataFrame from the collected response data
        response_df = pd.DataFrame(response_data, columns=["HCC ID", "Vendor ID", "Status", "CVC ID", "Error Message"])

        # Save the DataFrame to a CSV file
         # Append to the CSV file after each row or batch of rows
        response_df.to_csv(summary_file, mode='a', header=not summary_exists, index=False)
        
        # Clear the response_data to avoid duplicates in memory (does not affect log file)
        response_data.clear()

        # Update the existence of the summary file after the first write
        summary_exists = True
        
         # Increment the rows processed counter
        rows_processed += 1

        # Print the number of rows processed so far
        print(f"Rows processed in this set: {rows_processed}")


print(f"All requests for {current_env} have been sent and responses have been written to '{log_file}'.")
