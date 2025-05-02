import os
import requests
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

current_env = os.getenv('ENV', 'UAT') 

environments = {
    "Prod": {
        "url": "https://gchp-con-pr01.gchp.local:9193/connector/services/v4/ClaimReprocessing",
    },
    "UAT": {
        "url": "https://gchp-con-ut01.gchp.local:9193/connector/services/v4/ClaimReprocessing",
    },
}

if current_env not in environments:
    raise ValueError(f"Unknown environment: {current_env}")

config = environments[current_env]
url = config["url"]

excel_file = 'Input_File'
workbook = load_workbook(filename=excel_file)
sheet = workbook.active


with open(log_file, 'w') as response_file:
    
    for row in sheet.iter_rows(min_row=2, values_only=True):  
        claim_id = row[0]
        member_id = row[1]

        if len(str(member_id)) == 11:
            member_id = member_id[:9]

        soap_body = f"""
        <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:req="http://www.healthedge.com/connector/schema/claimreprocessing/request">
           <soapenv:Header/>
           <soapenv:Body>
              <req:claimReprocessingRequest>
                 <ClaimDetailsRequest>
                    <previewMode>false</previewMode>
                    <forceReprocessing>true</forceReprocessing>
                    <rematch>true</rematch>
                    <claimId>{claim_id}</claimId>
                    <header>
                       <member>
                          <memberId>{member_id}</memberId>
                       </member>
                    </header>
                    <auditLog>
                       <reason>Claim corrected with incorrect member Id</reason>
                       <comments>Claim corrected</comments>
                    </auditLog>
                 </ClaimDetailsRequest>
              </req:claimReprocessingRequest>
           </soapenv:Body>
        </soapenv:Envelope>
        """

        headers = {'Content-Type': 'text/xml'}
        response = requests.post(url, data=soap_body, headers=headers, verify=False)

        response_file.write(f"Response for Claim ID {claim_id} / Member ID {member_id}:\n")
        response_file.write(response.text + "\n\n")

print(f"All requests for {current_env} have been sent and responses have been written to '{log_file}'.")
