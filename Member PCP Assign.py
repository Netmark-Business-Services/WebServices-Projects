from datetime import datetime
import os
import requests
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from requests.auth import HTTPBasicAuth
import pandas as pd

# Load the environment variable
current_env = os.getenv('ENV', 'env')  # Default to 'Dev' if the environment variable is not set

# Define environment configurations
environments = {
    "UAT": {
        "url": "https://gchp-con-ut01.gchp.local:9193/connector/services/v4/EnrollmentSparse",
        "member_url": "https://gchp-con-ut01.gchp.local:9193/connector/services/v4/MembershipSparseLookup",

    },
    "Prod": {
        "url": "https://gchp-con-pr01.gchp.local:9193/connector/services/v4/EnrollmentSparse",
        "member_url": "https://gchp-con-ut01.gchp.local:9193/connector/services/v4/MembershipSparseLookup",

    },
    "PreProd": {
        "url": "https://gchp-con-pp01.gchp.local:9193/connector/services/v4/EnrollmentSparse",

    },
    "Dev": {
        "url": "https://gchp-con-dv01.gchp.local:9193/connector/services/v4/EnrollmentSparse",

    }
}

# Ensure the specified environment is valid
if current_env not in environments:
    raise ValueError(f"Unknown environment: {current_env}")

# Get the configuration for the current environment
config = environments[current_env]
url = config["url"]


# Load the Excel file
excel_file = ''
# Check if the file exists
if not os.path.exists(excel_file):
    print(f"File not found: {excel_file}")
    raise FileNotFoundError(f"The file does not exist in the specified path.")

workbook = load_workbook(filename=excel_file)
sheet = workbook.active

# Initialize a list to collect response data for the summary
response_data = []

summary_exists = os.path.exists(summary_file)wq
with open(log_file, 'w') as response_file:

    for row in sheet.iter_rows(min_row=min_row, max_row=max_row,  values_only=True):
        hcc_id = row[0]
        vendor_id = row[1]
        pcp_name = row[3]
        start_date = row[4]

        print(f"HCC ID: {hcc_id}, Vendor ID: {vendor_id}, PCP Name: {pcp_name}, Start Date: {start_date}")


        if isinstance(start_date, str):
            start_date = datetime.strptime(start_date, '%Y-%m-%d')  # Adjust format if necessary
        start_date_converted = start_date.strftime('%Y-%m-%d')

        vendor_id = clean_vendor_id(vendor_id)


        soap_body = f"""
        <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:enr="http://www.healthedge.com/connector/schema/enrollmentsparse">
            <soapenv:Body>
                <enr:enrollment>
                    <asOfDate>Date</asOfDate>
                    <sendToWorkBasketIfExceptionsPresent>false</sendToWorkBasketIfExceptionsPresent>
                    <member>
                        <maintenanceTypeCode>CHANGE</maintenanceTypeCode>
                        <hccIdentifier>{hcc_id}</hccIdentifier>
                        <memberMatchData>
                            <definitionName>MemberMatchDefinition</definitionName>
                            <id>{hcc_id}</id>
                        </memberMatchData>
                        <providerSelections>
                            <providerSelection>
                                <providerRoleType>PCP</providerRoleType>
                                <providerDateRanges>
                                    <startDate>{start_date_converted}</startDate>
                                    <endDate>Date</endDate>
                                    <providerMatch>
                                        <supplierLocation>
                                            <hccIdentificationNumber>{vendor_id}</hccIdentificationNumber>
                                        </supplierLocation>
                                    </providerMatch>
                                    <pcpAutoAssigned>false</pcpAutoAssigned>
                                </providerDateRanges>
                            </providerSelection>
                        </providerSelections>
                    </member>         
                </enr:enrollment>
            </soapenv:Body>
        </soapenv:Envelope>
        """

        # Send the SOAP request with basic authentication
        headers = {'Content-Type': 'text/xml'}
        response = requests.post(url, data=soap_body, headers=headers, verify=False)

        # Write the response to the file
        response_file.write(f"Response for HCC ID {hcc_id}:\n")
        response_file.write(response.text + "\n\n")

        # Print the response (for debugging)
        print(f"Response for HCC ID {hcc_id} in {current_env}:")
        print(response.text)

       # Parse the response to extract needed information
        root = ET.fromstring(response.text)

        status = root.find(".//status")
        status_text = status.text if status is not None else 'Unknown'

        cvc_id = root.find(".//cvcId")
        cvc_id_text = cvc_id.text if cvc_id is not None else 'Unknown'

        error_message = root.find(".//messageDescription")
        error_message_text = error_message.text if error_message is not None else 'None'

        # Collect response data
        response_data.append([hcc_id, vendor_id, status_text, cvc_id_text, error_message_text])
        
        # Create a DataFrame from the collected response data
        response_df = pd.DataFrame(response_data, columns=["HCC ID", "Vendor ID", "Status", "CVC ID", "Error Message"])

        # Save the DataFrame to a CSV file
         # Append to the CSV file after each row or batch of rows
        response_df.to_csv(summary_file, mode='a', header=not summary_exists, index=False)
        
        # Clear the response_data to avoid duplicates in memory (does not affect log file)
        response_data.clear()

        # Update the existence of the summary file after the first write
        summary_exists = True
        
         # Increment the rows processed counter
        rows_processed += 1

