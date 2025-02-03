import os
import requests
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from requests.auth import HTTPBasicAuth
import csv

# Load the environment variable
current_env = os.getenv('ENV', 'Prod')  # Default to 'Prod' if the environment variable is not set

# Define environment configurations
environments = {
    "Prod": {
        "url": "https://gchp-lb-pr01.gchp.local:5559/connector/services/v4/ClaimStatusLookup",
        "username": "username",
        "password": "password",
        "log_file": "C:\\Test complete\\Netmark\\Projects\\Procedure Code Modifier\\Updating PROD ENV\\other CPT codes Change\\line 4\\Claim-Status-Prod-Responses.txt",
        "csv_file": "C:\\Test complete\\Netmark\\Projects\\Procedure Code Modifier\\Updating PROD ENV\\other CPT codes Change\\line 4\\Claim-Status-Prod-Responses.csv",
    },
    # Add other environments if needed
}

# Ensure the specified environment is valid
if current_env not in environments:
    raise ValueError(f"Unknown environment: {current_env}")

# Get the configuration for the current environment
config = environments[current_env]
url = config["url"]
username = config["username"]
password = config["password"]
log_file = config["log_file"]
csv_file = config["csv_file"]

# Load the Excel file
excel_file = "C:\\Test complete\\Netmark\\Projects\\Procedure Code Modifier\\Demo Presentation\\claim id's for claim line 2.xlsx"  # Change this to your Excel file path

workbook = load_workbook(filename=excel_file)
sheet = workbook.active

# Open files to write the responses and CSV summary
with open(log_file, 'w') as response_file, open(csv_file, 'w', newline='') as csvfile:
    # Set up the CSV writer
    csv_writer = csv.writer(csvfile)
    # Write the header for the CSV file
    csv_writer.writerow(['HCC ID'] + [f'Service Code {i+1}' for i in range(20)])  # Add headers for up to 20 service codes

    # Iterate through the rows in the Excel file, assuming HCC IDs are in the first column
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
        hcc_id = row[0]  # HCC ID is in the first column

        # Create the SOAP request body
        soap_body = f"""
        <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:stat="http://www.healthedge.com/connector/schema/claim/status">
           <soapenv:Header/>
           <soapenv:Body>
              <stat:originalEDI837AttachmentReferenceLookupCriteria>
                 <hccClaimNumber>{hcc_id}</hccClaimNumber>
              </stat:originalEDI837AttachmentReferenceLookupCriteria>
           </soapenv:Body>
        </soapenv:Envelope>
        """

        # Send the SOAP request with basic authentication
        headers = {'Content-Type': 'text/xml'}
        response = requests.post(url, data=soap_body, headers=headers, auth=HTTPBasicAuth(username, password), verify=False)

        # Write the response to the Text file
        response_file.write(f"Response for HCC ID {hcc_id}:\n")
        response_file.write(response.text + "\n\n")

        # Parse the response to extract CPT codes
        try:
            root = ET.fromstring(response.content)

            # Extract the document element from the response
            document = root.find(".//document")
            service_codes = []

            if document is not None and document.text:
                # The document text is a CDATA containing XML, so we need to parse it as well
                internal_root = ET.fromstring(document.text)

                # Extract service codes for professional claims
                service_line_items = internal_root.findall(".//serviceLineItem")
                for service_line in service_line_items:
                    service_code = service_line.find("serviceCode").text if service_line.find("serviceCode") is not None else 'No service code'
                    service_codes.append(service_code)

                # Extract service codes for institutional claims
                institutional_service_lines = internal_root.findall(".//institutionalServiceLineItem/serviceLine")
                for service_line in institutional_service_lines:
                    service_code = service_line.find("serviceCode").text if service_line.find("serviceCode") is not None else 'No service code'
                    service_codes.append(service_code)

                # Write the HCC ID and Service Codes to the CSV file
                print(f"\nSuccess: HCC ID {hcc_id} - Successfully retrieved {len(service_codes)} service codes.\n")
                csv_writer.writerow([hcc_id] + service_codes + [''] * (20 - len(service_codes)))  # Fill empty columns with ''

            else:
                print(f"No document found for HCC ID {hcc_id}")
                csv_writer.writerow([hcc_id] + ['No document found'] + [''] * 19)

        except ET.ParseError as e:
            print(f"Error parsing XML for HCC ID {hcc_id}: {e}")
            csv_writer.writerow([hcc_id, f'Error parsing XML: {e}'] + [''] * 19)

print(f"\nAll requests for {current_env} have been sent and responses have been written to '{log_file}'.\n")
print(f"\nSummary data has been saved to '{csv_file}'.\n")
