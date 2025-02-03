import os
import requests
import xml.etree.ElementTree as ET
from openpyxl import load_workbook, Workbook
from requests.auth import HTTPBasicAuth
import warnings
from urllib3.exceptions import InsecureRequestWarning

# Suppress only the InsecureRequestWarning from urllib3, but we will log it without considering it a failure
warnings.filterwarnings('ignore', category=InsecureRequestWarning)

# Load the environment variable
current_env = os.getenv('ENV', 'Prod')  # Default to 'Prod' if not set

# Define environment configurations
environments = {
    "Prod": {
        "url": "https://gchp-lb-pr01.gchp.local:5559/connector/services/v4/ClaimReprocessing",
        "username": "username",
        "password": "password",
        "log_file": "C:\\Test complete\\Netmark\\Projects\\Procedure Code Modifier\\Updating PROD ENV\\Claim-Reprocessing-Prod-Responses.txt",
    },
}

# Ensure the specified environment is valid
if current_env not in environments:
    raise ValueError(f"Unknown environment: {current_env}")

# Get the configuration for the current environment
config = environments[current_env]
url = config["url"]
username = config["username"]
password = config["password"]
log_file = config["log_file"]

# Update file paths as per your requirement
input_excel_file =     "C:\\Test complete\\Netmark\\Projects\\Procedure Code Modifier\\Demo Presentation\\re-processed claims\\CPT codes for claim line 1.xlsx"
completed_excel_file = "C:\\Test complete\\Netmark\\Projects\\Procedure Code Modifier\\Demo Presentation\\re-processed claims\\CompletedClaims.xlsx.xlsx"
error_excel_file =     "C:\\Test complete\\Netmark\\Projects\\Procedure Code Modifier\\Demo Presentation\\re-processed claims\\FailedClaims.xlsx"

# Load the Excel file with claim data (Sheet 1)
workbook = load_workbook(filename=input_excel_file)
sheet = workbook['Sheet1']  # Assuming the sheet name is "Sheet 1"

# Create workbooks for logging completed and failed claims
completed_wb = Workbook()
completed_ws = completed_wb.active
completed_ws.append(["Claim ID", "Status"])

error_wb = Workbook()
error_ws = error_wb.active
error_ws.append(["Claim ID", "Error Type", "Error Message"])

# Initialize a counter
claim_count = 0

# Open a file to write the responses
with open(log_file, 'w') as response_file:
    # Iterate through the rows in the Excel file, assuming claimId and procedureCode are in the first two columns
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
        claim_count += 1  # Increment the claim counter
        claim_id = row[0]  # Claim ID
        procedure_code = row[1]  # Procedure Code

        # Create the SOAP request body
        soap_body = f"""
        <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:req="http://www.healthedge.com/connector/schema/claimreprocessing/request">
           <soapenv:Header/>
           <soapenv:Body>
              <req:claimReprocessingRequest>
                 <ClaimDetailsRequest>
                    <previewMode>false</previewMode>
                    <forceReprocessing>false</forceReprocessing>
                    <rematch>true</rematch>
                    <claimId>{claim_id}</claimId>
                    <header>
                    </header>
                    <lines>
                       <lineDetails>
                          <hccOriginalClaimLineNumber>1</hccOriginalClaimLineNumber>
                          <procedureCode>{procedure_code}</procedureCode>
                       </lineDetails>
                    </lines>
                 </ClaimDetailsRequest>
              </req:claimReprocessingRequest>
           </soapenv:Body>
        </soapenv:Envelope>
        """

        # Send the SOAP request with basic authentication
        headers = {'Content-Type': 'text/xml'}
        try:
            # Sending the request and ignoring InsecureRequestWarning
            response = requests.post(url, data=soap_body, headers=headers, auth=HTTPBasicAuth(username, password), verify=False)
            response_file.write(f"Response for Claim ID {claim_id}:\n")
            response_file.write(response.text + "\n\n")

            # Parse the response XML
            root = ET.fromstring(response.content)

            # Check if the response contains success or 'no change'
            status_tag = root.find(".//status")
            if status_tag is not None and status_tag.text == "SUCCESS":
                completed_ws.append([claim_id, "Updated"])
                print(f"{claim_count}. Claim ID {claim_id}: Updated")
            elif status_tag is not None and status_tag.text.lower() == "no change":
                completed_ws.append([claim_id, "No Change"])
                print(f"{claim_count}. Claim ID {claim_id}: No Change")
            elif status_tag is not None and status_tag.text.lower() == "error":
                # If there is an error status, extract errorType and message
                error_type = root.find(".//errorType").text if root.find(".//errorType") is not None else "Unknown error type"
                error_message = root.find(".//message").text if root.find(".//message") is not None else "No error message found"
                error_ws.append([claim_id, error_type, error_message])
                print(f"{claim_count}. Claim ID {claim_id}: Failed with error type - {error_type}, message - {error_message}")
            else:
                # Fallback error message in case of any other issues
                error_message = root.find(".//error").text if root.find(".//error") is not None else "Error not found in response"
                error_ws.append([claim_id, "Unknown", error_message])
                print(f"{claim_count}. Claim ID {claim_id}: Failed with error - {error_message}")

        except requests.exceptions.RequestException as e:
            # Check if the error is InsecureRequestWarning, if so, log it but don't consider it a failure
            if "InsecureRequestWarning" in str(e):
                response_file.write(f"Warning for Claim ID {claim_id}: {e}\n\n")
                print(f"{claim_count}. InsecureRequestWarning for Claim ID {claim_id}: {e}")  # Print but don't log as failed
            else:
                # Log any other exceptions in the error file
                error_ws.append([claim_id, "Request Exception", str(e)])
                response_file.write(f"Error for Claim ID {claim_id}: {e}\n\n")
                print(f"{claim_count}. Claim ID {claim_id}: Failed with exception - {str(e)}")

# Save the completed and error Excel files
completed_wb.save(completed_excel_file)
error_wb.save(error_excel_file)

print(f"All requests for {current_env} have been sent and responses have been written to '{log_file}'.\n")
print(f"Completed claims saved to '{completed_excel_file}'.\n")
print(f"Failed claims saved to '{error_excel_file}'.")
