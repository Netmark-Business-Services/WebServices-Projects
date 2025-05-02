import os
import pandas as pd
import requests
from requests.auth import HTTPBasicAuth
import time
import logging
from datetime import datetime
import re
import xml.etree.ElementTree as ET
import openpyxl


# Determine the base directory of the script
base_dir = os.path.dirname(__file__)  # Base directory where MTM.py is located

# Load the Excel file with Supplier data using a relative path
supplier_data_file = os.path.join(base_dir, 'Spreadsheet Claims','Fresh and Fabulous-FAF','F&F claims resubmit 4_15_25.xlsx')

# Extract the supplier name from the file path
file_name = os.path.splitext(os.path.basename(supplier_data_file))[0]

# Ensure the logs directory exists
os.makedirs(os.path.join(base_dir, 'logs'), exist_ok=True)

# Set up logging to use the relative path for logs, with FileHandler
log_filename = os.path.join(base_dir, 'logs', f'{file_name}_soap_client_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename),  # Set up a file handler with the specified filename
        logging.StreamHandler()  # Also log to the console
    ]
)
logger = logging.getLogger(__name__)


# Define environment configurations with relative log paths
environments = {
    "UAT": {
        "url": "https://gchp-con-ut01.gchp.local:9193/connector/services/v4/ProfessionalClaim",
        "status_url": "https://gchp-con-ut01.gchp.local:9193/connector/services/v4/ClaimStatusLookup",  # Claim status URL
        "username": "",
        "password": "",
        "req_log_file": os.path.join(base_dir, 'logs',f"{file_name} Claim-Submit-UAT-Requests.txt"),
        "res_log_file": os.path.join(base_dir, 'logs', f"{file_name} Claim-Submit-UAT-Responses.txt"),
    }
}

# Address Parsing Function
def parse_address(address):
    """
    Parse an address string into street address, city name, state code, and postal code.
    """
    #pattern = r'(.+),\s*([A-Za-z\s]+),\s*([A-Z]{2}),\s*(\d{5})'
    #pattern = r"^(\d+\s[^,#]+(?:,\sL\d+| # \d+)?),?\s([^,]+),\s(\w{2}),\s(\d{5})$" # Pattern to handle Ventura county AAA (Tested)
    #pattern = r"^(\d+\s[^,]+,\sSTE\s\d+),\s([^,]+),\s(\w{2})\s(\d{5})$" # Pattern to handle International Elder Care-IHECS (Tested)
    pattern = r'(.+?)[,\s]+([A-Za-z\s]+)[,\s]+([A-Z]{2})\s+(\d{5})$' # Pattern to handle Divine Agape Health Care Agency, The Medical Kitchen-TMK, Fresh and Fabulous-FAF
    #pattern = r'(.+?),\s*([A-Za-z\s]+),\s*([A-Z]{2})[.,]?\s+(\d{5})$' # Updated pattern to handle more variations in address format (Ventura Harbor - Waters Edge - TAT, International Elder Care-IHECS)
    match = re.search(pattern, address)
    print(match)
    if match:
        street_address = match.group(1).strip()
        city_name = match.group(2).strip()
        state_code = match.group(3).strip()
        postal_code = match.group(4).strip()
        return street_address, city_name, state_code, postal_code
    else:
        return None, None, None, None

# SOAP Client Class
class SOAPClient:
    def __init__(self, environment="Prod"):
        """Initialize the SOAP client with environment configuration"""
        self.current_env = environment
        self.config = environments[environment]
        self.url = self.config["url"]
        self.status_url = self.config["status_url"]
        self.username = self.config["username"]
        self.password = self.config["password"]
        self.req_log_file = self.config["req_log_file"]
        self.res_log_file = self.config["res_log_file"]

        # Ensure log directory exists
        os.makedirs(os.path.dirname(self.req_log_file), exist_ok=True)

        # Set up session with basic authentication
        self.session = requests.Session()
        self.session.auth = HTTPBasicAuth(self.username, self.password)
        self.session.verify = False

        # Headers from WSDL specification
        self.headers = {
            'Content-Type': 'text/xml;charset=UTF-8',
            'SOAPAction': 'http://healthedge.com/submit'
        }

        # Headers from WSDL specification
        self.claim_status_headers = {
            'Content-Type': 'text/xml;charset=UTF-8',
            'SOAPAction': 'http://healthedge.com/getAll'
        }


    def normalize_value(self, value):
        """Normalize values for consistent format"""
        if isinstance(value, str):
            if value.lower() in ['yes', 'y']:
                return 'Yes'
            if value.lower() in ['no', 'n']:
                return 'No'
        return value

    def normalize_gender_value(self, value):
        # Normalize values for consistent format
        if isinstance(value, str):
            value = value.strip()  # Remove leading and trailing spaces
            if value.lower() == 'm':
                return 'M'
            if value.lower() == 'f':
                return 'F'
        return value


    def create_soap_request_for_member(self, grouped_member_data):
        """
        Generate the SOAP request XML for a given member's grouped data.
        """
        try:
            first_row = grouped_member_data.iloc[0]
            member_id = first_row['Member ID']
            member_name = f"{first_row['Member First name']} {first_row['Member Last Name']}"
            member_dob = pd.to_datetime(first_row['Member DOB']).strftime('%Y-%m-%d')

            # Normalize values for consistency
            release_auth = self.normalize_value(first_row['Release Authorization Signature'])
            insured_sig = self.normalize_value(first_row['Insureds Signature'])
            assignment_acceptance = self.normalize_value(first_row['Accept Assignment'])
            benefit_assignment = self.normalize_value(first_row['Benefit Assignment'])

            #Normalize values for gender consistency
            Gender = self.normalize_gender_value(first_row['Sex'])

            # Parse Billing Address
            supplier_street, supplier_city, supplier_state, supplier_zip = parse_address(first_row['Billing Address'])

            # Parse Service Location Address
            facility_street, facility_city, facility_state, facility_zip = parse_address(first_row['Service Location Address'])

            # Initialize the SOAP body
            soap_body = f"""<?xml version="1.0" encoding="UTF-8"?>
<soapenv:Envelope
    xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/"
    xmlns:prof="http://www.healthedge.com/connector/schema/claim/professional"
    xmlns:resp="http://www.healthedge.com/connector/schema/claimresponse">
   <soapenv:Header/>
   <soapenv:Body>
      <prof:professionalClaim resourceURI="http://www.healthedge.com/connector/schema/claim" version="1.0">
         <receiptDate>{pd.to_datetime(first_row['Receipt Date']).strftime('%Y-%m-%d')}</receiptDate>
         <entryDate>{time.strftime("%Y-%m-%dT%H:%M:%S")}</entryDate>
         <cleanClaimDate>{pd.to_datetime(first_row['Clean Claim Date']).strftime('%Y-%m-%d')}</cleanClaimDate>
         <frequencyCode>1</frequencyCode>

         <subscriberInformation>
            <subscriberIdentificationNumber>{member_id}</subscriberIdentificationNumber>
            <subscriberName>{member_name}</subscriberName>
            <lastName>{first_row['Member Last Name']}</lastName>
            <firstName>{first_row['Member First name']}</firstName>
            <genderCode>{Gender}</genderCode>
            <dateOfBirth>{member_dob}</dateOfBirth>
            <benefitAssignment>{benefit_assignment}</benefitAssignment>
         </subscriberInformation>

         <memberInformation>
            <memberIdentificationNumber>{member_id}</memberIdentificationNumber>
            <memberName>{member_name}</memberName>
            <lastName>{first_row['Member Last Name']}</lastName>
            <firstName>{first_row['Member First name']}</firstName>
            <genderCode>{Gender}</genderCode>
            <dateOfBirth>{member_dob}</dateOfBirth>
            <relationshipToSubscriberCode>{first_row['Patient Relationship to Insured']}</relationshipToSubscriberCode>
            <accountNumber>{first_row['Patient Account Number']}</accountNumber>
         </memberInformation>

         <memberAuthorization>
            <releaseAuthorization>{release_auth}</releaseAuthorization>
            <insuredSignature>{insured_sig}</insuredSignature>
         </memberAuthorization>

         <supplierInformationList>
            <supplierInformation>
               <supplierBillingName>{first_row['Billing Name'].replace('&', '&amp;')}</supplierBillingName>
               <taxIdentificationNumber>{first_row['TIN']}</taxIdentificationNumber>
               <assignmentAcceptance>{assignment_acceptance}</assignmentAcceptance>
               <streetAddress>{supplier_street}</streetAddress>
               <cityName>{supplier_city}</cityName>
               <stateCode>{supplier_state}</stateCode>
               <postalCode>{supplier_zip}</postalCode>
               <countryCode>US</countryCode>
               <npi>{first_row['Billing NPI']}</npi>
            </supplierInformation>           
         </supplierInformationList>

         <diagnosisCodes>
            <diagnosisCode>{first_row['Diagnosis Code']}</diagnosisCode>
         </diagnosisCodes>

         <renderingFacility>
            <facilityName>{first_row['Service Location Name'].replace('&', '&amp;')}</facilityName>
            <streetAddress>{facility_street}</streetAddress>
            <cityName>{facility_city}</cityName>
            <stateCode>{facility_state}</stateCode>
            <postalCode>{facility_zip}</postalCode>
            <countryCode>US</countryCode>
            <npi>{first_row['Service location NPI']}</npi>
         </renderingFacility>

         <totalPaymentDue>
            <chargedAmount>{first_row['Total Charge']}</chargedAmount>
         </totalPaymentDue>

         <serviceLineItems>"""
            # Add each service line item
            original_line_number = 1  # Start numbering from 1
            for index, item in grouped_member_data.iterrows():
                start_date = pd.to_datetime(item['DOS From']).strftime('%Y-%m-%d')
                end_date = pd.to_datetime(item['DOS To']).strftime('%Y-%m-%d')
                
                soap_body += f"""
            <serviceLineItem>
            
               <originalLineNumber>{original_line_number}</originalLineNumber>
               <startDate>{start_date}</startDate>
               <endDate>{end_date}</endDate>
               <placeOfServiceCode>{item['POS']}</placeOfServiceCode>
               <serviceCode>{item['Procedure Code']}</serviceCode>
               <serviceFee>{item['Line Charges']}</serviceFee>
               <serviceUnitCount>{item['Units']}</serviceUnitCount>
               <renderingProviderNPI>{item['Rendering Provider NPI']}</renderingProviderNPI>
               <modifierList>
                  <modifier>{item['Modifier']}</modifier>
               </modifierList>
               <diagnosisCodePointers>
                  <diagnosisCodePointer>{item['Diagnosis Pointer']}</diagnosisCodePointer>
               </diagnosisCodePointers>
            </serviceLineItem>"""
                original_line_number += 1  # Increment for the next item

            soap_body += """
         </serviceLineItems>
      </prof:professionalClaim>
   </soapenv:Body>
</soapenv:Envelope>"""

            return soap_body.strip()

        except Exception as e:
            logger.error(f"Error creating SOAP request: {str(e)}")
            raise

    def submit_claims_from_data(self, test_data_df):
        """
        Submit SOAP requests for each member group in the test data.
        """
        grouped_claims = test_data_df.groupby('Member ID')
       
        with open(self.req_log_file, 'w', encoding='utf-8') as request_file:
            with open(self.res_log_file, 'w', encoding='utf-8') as response_file:
                for member_id, group in grouped_claims:
                    try:
                        soap_request = self.create_soap_request_for_member(group)
                        if soap_request is None:
                            continue

                        member_info = f"{group.iloc[0]['Member First name']} {group.iloc[0]['Member Last Name']} ({member_id})"
                       
                        # Log request with visual separators
                        logger.info(f"\n{'='*80}\nProcessing request for {member_info}\n{'='*80}")
                       
                        request_file.write(f"\n{'='*80}\nRequest for Member {member_info}\n{'='*80}\n")
                        request_file.write(f"Headers: {self.headers}\n")
                        request_file.write(f"Request Body:\n{soap_request}\n\n")
                       
                        # Send request
                        response = self.session.post(
                            self.url,
                            data=soap_request.encode('utf-8'),
                            headers=self.headers
                        )
                       
                        # Log response details
                        response_info = (
                            f"\nResponse for {member_info}:\n"
                            f"Status Code: {response.status_code}\n"
                            f"Response Headers: {dict(response.headers)}\n"
                            f"Response Body:\n{response.text}"
                        )
                       
                        logger.info(response_info)
                        response_file.write(f"{response_info}\n\n")
                       
                        if response.status_code != 200:
                            logger.warning(f"Non-200 status code ({response.status_code}) for member {member_info}")
                       
                    except Exception as e:
                        error_msg = f"Error processing member {member_id}: {str(e)}"
                        logger.error(error_msg)
                        response_file.write(f"{error_msg}\n\n")
                    time.sleep(2)

    
    def get_claim_status(self, claim_number):
        """
        Send a request to retrieve the status of the given claim number and parse the claim state from the response.
        """
        status_request_body = f"""<?xml version="1.0" encoding="UTF-8"?>
<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:stat="http://www.healthedge.com/connector/schema/claim/status">
   <soapenv:Header/>
   <soapenv:Body>
      <stat:claimStatusLookupCriteria resourceURI="http://healthedge.com/getAll" version="8.1">
         <hccClaimNumber>{claim_number}</hccClaimNumber>
      </stat:claimStatusLookupCriteria>
   </soapenv:Body>
</soapenv:Envelope>"""

        # Paths for logging request and response
        status_req_log_file = os.path.join(base_dir, 'logs', f"{file_name} ClaimStatusLookup-Requests.txt")
        status_res_log_file = os.path.join(base_dir, 'logs', f"{file_name} ClaimStatusLookup-Responses.txt")

        # Log the request
        with open(status_req_log_file, 'a', encoding='utf-8') as req_log:
            req_log.write(f"\n{'='*80}\nRequest for Claim Status - Claim Number: {claim_number}\n{'='*80}\n")
            req_log.write(f"Request Body:\n{status_request_body}\n\n")

        # Send the request to the status URL
        response = self.session.post(
            self.status_url,
            data=status_request_body,
            headers=self.claim_status_headers
        )

        # Log the response
        with open(status_res_log_file, 'a', encoding='utf-8') as res_log:
            res_log.write(f"\n{'='*80}\nResponse for Claim Status - Claim Number: {claim_number}\n{'='*80}\n")
            res_log.write(f"Status Code: {response.status_code}\n")
            res_log.write(f"Response Body:\n{response.text}\n\n")

        # Parse the claim state directly
        if response.status_code == 200:
            # Extract <claimState> from the response
            claim_state_match = re.search(r'<claimState>(.*?)</claimState>', response.text)

            if claim_state_match:
                claim_state = claim_state_match.group(1)
                logger.info(f"Claim {claim_number} Status: {claim_state}")
                return claim_state
            else:
                logger.warning(f"Claim state not found in the response for claim {claim_number}")
                return "Unknown"
        else:
            logger.error(f"Failed to retrieve status for claim {claim_number}. Status code: {response.status_code}")
            return "Error"

    def update_claim_status_in_excel(self, success_output_file):
        """
        Read 'Claim ID' values from the success Excel file, get the claim status,
        and write it in a new column 'Claim Status' next to each 'Claim ID'.
        """
        try:
            # Load the existing Excel file
            wb = openpyxl.load_workbook(success_output_file)
            ws = wb.active

            # Check for the "Claim ID" column and add "Claim Status" if not present
            if "Claim Status" not in [cell.value for cell in ws[1]]:
                ws.cell(row=1, column=ws.max_column + 1, value="Claim Status")

            # Find the columns for 'Claim ID' and 'Claim Status'
            claim_id_col = None
            status_col = None
            for idx, cell in enumerate(ws[1], 1):  # enumerate from 1 for openpyxl 1-indexing
                if cell.value == "Claim ID":
                    claim_id_col = idx
                elif cell.value == "Claim Status":
                    status_col = idx

            # Check that both columns were found
            if not claim_id_col or not status_col:
                logger.error("Missing required columns 'Claim ID' or 'Claim Status' in Excel.")
                return

            # Process each claim ID and update the status
            for row in range(2, ws.max_row + 1):  # Starting from row 2 to skip the header
                claim_id = ws.cell(row=row, column=claim_id_col).value
                if claim_id:
                    # Get claim status from the SOAP request
                    claim_status = self.get_claim_status(claim_id)
                    
                    # Write the status next to the Claim ID
                    ws.cell(row=row, column=status_col, value=claim_status)

            # Save the updated workbook
            wb.save(success_output_file)
            logger.info(f"Updated claim statuses written to {success_output_file}")

        except Exception as e:
            logger.error(f"Error updating claim statuses in Excel: {str(e)}")
            raise


    def parse_and_write_response_to_excel(self, success_output_file, error_output_file):
        """
        Parse the response log file to extract claim numbers and statuses,
        and write the results to separate Excel files for successful and unsuccessful claims.
        """
        try:
            successful_claims = []
            unsuccessful_claims = []
            current_member_id = None
            current_error_member_id = None
            error_message = None

            with open(self.res_log_file, 'r', encoding='utf-8') as file:
                content = file.read().strip().splitlines()

                for index, line in enumerate(content):
                    claim_number = None
                    status = None

                    # Check for 'Status Code: 200' to determine success and capture member ID from previous line
                    if "Status Code: 200" in line:
                        previous_line = content[index - 1]
                        if "Response for" in previous_line and '(' in previous_line and ')' in previous_line:
                            match = re.search(r'\((.*?)\)', previous_line)
                            if match:
                                current_member_id = match.group(1)
                                error_message = None  # Reset error message for new request

                    # Check for 'Status Code: 500' to determine failure and capture error message
                    if "Status Code: 500" in line:
                        # Look for the 'errors' block in the subsequent lines
                        errors_block = []
                        for i in range(index + 1, len(content)):
                            if "errors:" in content[i]:
                                match = re.search(r"errors:\s*\[(.*?)\]\s*\.","\n".join(content[i:]), re.DOTALL)
                                if match:
                                    errors_block.append(match.group(1).strip())
                                break
                            elif "</soap:Fault>" in content[i]:  # Stop parsing if the fault block ends
                                break

                        if errors_block:
                            error_message = " | ".join(errors_block)  # Combine error messages if multiple found

                        # Capture Member ID for the failed response
                        previous_line = content[index - 1]
                        if "Response for" in previous_line and '(' in previous_line and ')' in previous_line:
                            match = re.search(r'\((.*?)\)', previous_line)
                            if match:
                                current_error_member_id = match.group(1)

                    # Capture error message if line contains "Error processing member"
                    if "Error processing member" in line:
                        match = re.search(r"Error processing member (\w+):", line)
                        if match:
                            current_error_member_id = match.group(1)
                            error_message = line.strip()  # Capture the entire error line as the error message

                    # Extract claim number and status from the SOAP response XML
                    if '<hccClaimNumber>' in line:
                        claim_number = re.search(r'<hccClaimNumber>(.*?)</hccClaimNumber>', line).group(1)
                    if '<status>' in line:
                        status = re.search(r'<status>(.*?)</status>', line).group(1)

                    # Add successful claims
                    if current_member_id and claim_number and status == "SUCCESS":
                        successful_claims.append({
                            "Member ID": current_member_id,
                            "Claim ID": claim_number,
                            "Claim Response Status": status
                        })
                        current_member_id = None

                    # Add unsuccessful claims with errors
                    elif current_error_member_id and error_message:
                        unsuccessful_claims.append({
                            "Member ID": current_error_member_id,
                            "Error": error_message
                        })
                        current_error_member_id = None
                        error_message = None

            # Write successful claims to Excel
            if successful_claims:
                df_successful = pd.DataFrame(successful_claims)
                df_successful.to_excel(success_output_file, index=False)
                logger.info(f"Successful claims written to Excel at {success_output_file}")

            # Write unsuccessful claims to Excel
            if unsuccessful_claims:
                df_unsuccessful = pd.DataFrame(unsuccessful_claims)
                df_unsuccessful.to_excel(error_output_file, index=False)
                logger.info(f"Unsuccessful claims written to Excel at {error_output_file}")
            else:
                logger.info("No unsuccessful claims to write.")

        except Exception as e:
            logger.error(f"Error parsing response log file: {str(e)}")
            raise


def main():
    try:
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        current_env = os.getenv('ENV', 'Prod')
        logger.info(f"Starting SOAP client in {current_env} environment")
       
        client = SOAPClient(environment=current_env)
       
        logger.info(f"Reading data from {supplier_data_file}")
        df_supplier_data = pd.read_excel(supplier_data_file, dtype=str)
       
        value = 'Yes'
        df_supplier_data['Release Authorization Signature'] = value
        df_supplier_data['Insureds Signature'] = value
        df_supplier_data['Accept Assignment'] = value
        df_supplier_data['Benefit Assignment'] = value

        date_value = '2025-04-16'
        df_supplier_data['Clean Claim Date'] = date_value
        df_supplier_data['Receipt Date'] = date_value


        df_supplier_data['TIN'] = df_supplier_data['TIN'].apply(
            lambda x: f"{x[:2]}-{x[2:]}" if pd.notnull(x) and len(x) == 9 and '-' not in x else x
        )

        print("Available columns:", df_supplier_data.columns.tolist())

       
        client.submit_claims_from_data(df_supplier_data)
        logger.info(f"Process completed. Check logs at:")
        logger.info(f"Requests: {client.req_log_file}")
        logger.info(f"Responses: {client.res_log_file}")

        # Define the paths for both output files
        success_output_file = os.path.join(base_dir, 'logs', f"(Claims Created) {file_name}.xlsx")
        error_output_file = os.path.join(base_dir, 'logs', f"(Claims Not Created) {file_name}.xlsx")

        # Pass both output files to the function
        client.parse_and_write_response_to_excel(success_output_file, error_output_file)
        logger.info(f"Claims responses written to Excel at {success_output_file}")
        logger.info(f"Error claims written to Excel at {error_output_file}")

        # Update claim status in the success Excel file
        client.update_claim_status_in_excel(success_output_file)
        logger.info(f"Claim statuses updated in {success_output_file}")

    except Exception as e:
        logger.error(f"Fatal error: {str(e)}")
        raise

if __name__ == "__main__":
    main()