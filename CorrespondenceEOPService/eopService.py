from datetime import datetime
import os
import requests
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from requests.auth import HTTPBasicAuth
import pandas as pd

# Load the environment variable
current_env = os.getenv('ENV', 'UAT')  # Default to 'Dev' if the environment variable is not set


# Define environment configurations
environments = {
    "UAT": {
        "url": "https://gchp-con-ut01.gchp.local:9193/connector/services/classic/CorrespondenceServiceStronglyTyped",
        "successLog_file": "logs/eop-correspondence-success", 
        "failureLog_file": "logs/eop-correspondence-failure",
        "summary_file": "logs/eop-correspondence-summary"  
    },
    "Prod": {
        "url": "https://gchp-con-pr01.gchp.local:9193/connector/services/classic/CorrespondenceServiceStronglyTyped",
        "log_file": "logs/eop-correspondence-responses",
        "summary_file": "logs/eop-correspondence-responses"
    },
    "PreProd": {
        "url": "https://gchp-con-pp01.gchp.local:9193/connector/services/classic/CorrespondenceServiceStronglyTyped",
        "log_file": "logs/eop-correspondence-responses",
        "summary_file": "logs/eop-correspondence-responses"
    }
}

# Ensure the specified environment is valid
if current_env not in environments:
    raise ValueError(f"Unknown environment: {current_env}")

# Get the configuration for the current environment
config = environments[current_env]
url = config["url"]
successLog_file = config["successLog_file"]
failureLog_file = config["failureLog_file"]
summary_file = config["summary_file"]

# Load the Excel file
excel_file = ''

workbook = load_workbook(filename=excel_file)
sheet = workbook.active

# Initialize lists to collect response data
success_log = []
failure_log = []
summary = []

# Initialize a counter for the number of rows processed
rows_processed = 0

for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
    supplierHccId = row[1]
    paymentId = row[2]

    print(f"SupplierHCCID: {supplierHccId}, PaymentID : {paymentId}")

     #Send the SOAP request
    soap_body = f"""
    <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:heal="http://healthedge.com">
    <soapenv:Header/>
    <soapenv:Body>
        <heal:createOrUpdateCorrespondence>
            <statusTimeList>
            <correspondenceStatus>t</correspondenceStatus>
            </statusTimeList>
            <description>Supplier EOP Letter</description>
            <definition>EOP manual</definition>
            <recipient>
                <hccId>{supplierHccId}</hccId>
                <recipientType>Supplier</recipientType>
            </recipient>
                <hccId>{paymentId}</hccId>
                <subjectType>Payment</subjectType>
        </heal:createOrUpdateCorrespondence>
    </soapenv:Body>
    </soapenv:Envelope>
        """

    headers = {'Content-Type': 'text/xml'}
    response = requests.post(url, data=soap_body, headers=headers, verify=False)

    if response.status_code == 200:
        root = ET.fromstring(response.text)
        status = root.find(".//status")
        status_text = status.text if status is not None else 'Success'

        success_log.append([supplierHccId, paymentId, "Success"])
        summary.append([supplierHccId, paymentId, "Success", response.text])
    else:
        failure_log.append([supplierHccId, paymentId, "Failure"])
        summary.append([supplierHccId, paymentId, "Failure", response.text])


# Save success and failure logs to CSV files
pd.DataFrame(success_log, columns=["HCC ID", "Payment ID", "Status"]).to_csv(successLog_file, index=False)
pd.DataFrame(failure_log, columns=["HCC ID", "Payment ID", "Status"]).to_csv(failureLog_file, index=False)
pd.DataFrame(summary, columns=["HCC ID", "Payment ID", "Status", "Message Description"]).to_csv(summary_file, index=False)

print(f"All requests for {current_env} have been processed.")
print(f"Successful responses saved to '{successLog_file}'.")
print(f"Failed responses saved to '{summary}'.")


# Print the starting row for the next set
next_starting_row = min_row + rows_processed
print(f"Next set should start from row: {next_starting_row}")











