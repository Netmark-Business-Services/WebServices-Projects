import os
import requests
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from requests.auth import HTTPBasicAuth
import csv


current_env = os.getenv('ENV', 'Prod') 

environments = {
    "Prod": {
        "url": "https://gchp-lb-pr01.gchp.local:5559/connector/services/v4/ClaimStatusLookup",
        "username": "username",
        "password": "password",
        "log_file": "Responses.txt",
        "csv_file": "Summary.csv",
    },
    "PreProd": {
        "url": "https://gchp-con-pp01.gchp.local:9193/connector/services/v4/EnrollmentSparse",
        "username": "username",
        "password": "password",
        "log_file": "Responses.txt",
        "csv_file": "Summary.csv",
    }, 
}


if current_env not in environments:
    raise ValueError(f"Unknown environment: {current_env}")

config = environments[current_env]
url = config["url"]
username = config["username"]
password = config["password"]
log_file = config["log_file"]
csv_file = config["csv_file"]

excel_file = 'input_File'  

workbook = load_workbook(filename=excel_file)
sheet = workbook.active

with open(log_file, 'w') as response_file, open(csv_file, 'w', newline='') as csvfile:
    csv_writer = csv.writer(csvfile)
    csv_writer.writerow(['HCC ID', 'Primary Diagnosis Code'] + [f'Other Diagnosis Code {i+1}' for i in range(11)])

    for row in sheet.iter_rows(min_row=2, values_only=True):
        hcc_id = row[0]  

        soap_body = f"""
        <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:stat="http://www.healthedge.com/connector/schema/claim/status">
           <soapenv:Header/>
           <soapenv:Body>
              <stat:originalEDI837AttachmentReferenceLookupCriteria>
                 <hccClaimNumber>{hcc_id}</hccClaimNumber>
              </stat:originalEDI837AttachmentReferenceLookupCriteria>
           </soapenv:Body>
        </soapenv:Envelope>
        """

        headers = {'Content-Type': 'text/xml'}
        response = requests.post(url, data=soap_body, headers=headers, auth=HTTPBasicAuth(username, password), verify=False)

        response_file.write(f"Response for HCC ID {hcc_id}:\n")
        response_file.write(response.text + "\n\n")

        try:
            root = ET.fromstring(response.content)

            document = root.find(".//document")
            if document is not None and document.text:
                internal_root = ET.fromstring(document.text)
                diagnosis_codes = internal_root.findall(".//diagnosisCode")

                if diagnosis_codes:
                    primary_diagnosis_code = diagnosis_codes[0].text
                    other_diagnosis_codes = [code.text for code in diagnosis_codes[1:12]]
                    csv_writer.writerow([hcc_id, primary_diagnosis_code] + other_diagnosis_codes)
                else:
                    print(f"No diagnosis codes found for HCC ID {hcc_id}")
                    csv_writer.writerow([hcc_id, 'No diagnosis codes found'] + [''] * 11)

            else:
                print(f"No document found for HCC ID {hcc_id}")
                csv_writer.writerow([hcc_id, 'No document found'] + [''] * 11) 

        except ET.ParseError as e:
            print(f"Error parsing XML for HCC ID {hcc_id}: {e}")
            csv_writer.writerow([hcc_id, f'Error parsing XML: {e}'] + [''] * 11) 

print(f"All requests for {current_env} have been sent and responses have been written to '{log_file}'.")
print(f"Summary data has been saved to '{csv_file}'.")
